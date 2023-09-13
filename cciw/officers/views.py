import contextlib
import enum
import json
from collections.abc import Callable, Iterable
from datetime import date, datetime
from functools import wraps
from typing import TypeAlias

import furl
import openpyxl
import pandas as pd
import pandas_highcharts.core
from django.conf import settings
from django.contrib import messages
from django.contrib.admin import site as admin_site
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.views import PasswordResetView
from django.core import signing
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied
from django.db.models import Prefetch
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.http.request import HttpRequest
from django.shortcuts import get_object_or_404
from django.template.defaultfilters import wordwrap
from django.template.response import TemplateResponse
from django.urls import reverse
from django.urls.resolvers import ResolverMatch, get_resolver
from django.utils import timezone
from django.views.decorators.cache import cache_control, never_cache
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_POST

from cciw.accounts.models import User
from cciw.bookings.models import Booking, Price, is_booking_open, most_recent_booking_year
from cciw.bookings.stats import get_booking_ages_stats, get_booking_progress_stats, get_booking_summary_stats
from cciw.bookings.utils import (
    addresses_for_mailing_list,
    camp_bookings_to_spreadsheet,
    camp_sharable_transport_details_to_spreadsheet,
    payments_to_spreadsheet,
    year_bookings_to_spreadsheet,
)
from cciw.cciwmain import common
from cciw.cciwmain.common import CampId
from cciw.cciwmain.decorators import json_response
from cciw.cciwmain.models import Camp
from cciw.cciwmain.utils import get_protected_download
from cciw.mail.lists import address_for_camp_officers, address_for_camp_slackers
from cciw.utils import xl
from cciw.utils.spreadsheet import ExcelBuilder, ExcelFromDataFrameBuilder
from cciw.utils.views import for_htmx, get_redirect_from_request, make_get_request, user_passes_test_improved

from . import create
from .applications import (
    application_rtf_filename,
    application_to_rtf,
    application_to_text,
    application_txt_filename,
    applications_for_camp,
    camps_for_application,
    invitations_for_application,
    thisyears_applications,
)
from .dbs import get_officers_with_dbs_info_for_camps
from .email import (
    make_ref_form_url,
    make_ref_form_url_hash,
    send_dbs_consent_alert_leaders_email,
    send_nag_by_officer,
    send_reference_request_email,
    send_request_for_dbs_form_email,
)
from .email_utils import formatted_email, send_mail_with_attachments
from .forms import (
    AdminReferenceForm,
    CciwPasswordResetForm,
    CorrectRefereeDetailsForm,
    CreateOfficerForm,
    DBSCheckForm,
    DbsConsentProblemForm,
    ReferenceForm,
    RequestDbsFormForm,
    SendNagByOfficerForm,
    SendReferenceRequestForm,
    UpdateOfficerForm,
)
from .models import (
    Application,
    CampRole,
    DBSActionLogType,
    DBSCheck,
    Invitation,
    OfficerList,
    Referee,
    Reference,
    ReferenceAction,
    add_officer_to_camp,
    add_previous_references,
    empty_reference,
    get_previous_references,
    remove_officer_from_camp,
)
from .stats import get_camp_officer_stats, get_camp_officer_stats_trend
from .utils import camp_serious_slacker_list, camp_slacker_list, officer_data_to_spreadsheet

EXPORT_PAYMENT_DATE_FORMAT = "%Y-%m-%d"

BOOKING_STATS_PREVIOUS_YEARS = 4


def _copy_application(application):
    new_obj = Application(id=None)
    for field in Application._meta.fields:
        if field.attname != "id":
            setattr(new_obj, field.attname, getattr(application, field.attname))
    new_obj.youth_work_declined = None
    new_obj.relevant_illness = None
    new_obj.crime_declaration = None
    new_obj.court_declaration = None
    new_obj.concern_declaration = None
    new_obj.allegation_declaration = None
    new_obj.dbs_check_consent = None
    new_obj.finished = False
    new_obj.date_saved = None
    new_obj.save()

    for old_ref, new_ref in zip(application.referees, new_obj.referees):
        for f in ["name", "address", "tel", "mobile", "email"]:
            setattr(new_ref, f, getattr(old_ref, f))
        new_ref.save()

    for q in application.qualifications.all():
        new_q = q.copy(application=new_obj)
        new_q.save()

    return new_obj


def any_passes(*funcs):
    def func(*args, **kwargs):
        for f in funcs:
            if f(*args, **kwargs):
                return True
        return False

    return func


camp_admin_required = user_passes_test_improved(lambda u: u.is_camp_admin)
dbs_officer_required = user_passes_test_improved(lambda u: u.is_dbs_officer)
dbs_officer_or_camp_admin_required = user_passes_test_improved(lambda u: u.is_dbs_officer or u.is_camp_admin)
booking_secretary_required = user_passes_test_improved(lambda u: u.is_booking_secretary)
booking_secretary_or_treasurer_required = user_passes_test_improved(
    any_passes(lambda u: u.is_booking_secretary, lambda u: u.is_treasurer)
)
cciw_secretary_required = user_passes_test_improved(lambda u: u.is_cciw_secretary)
cciw_secretary_or_booking_secretary_required = user_passes_test_improved(
    any_passes(lambda u: u.is_booking_secretary, lambda u: u.is_cciw_secretary)
)
secretary_or_committee_required = user_passes_test_improved(
    any_passes(lambda u: u.is_booking_secretary, lambda u: u.is_cciw_secretary, lambda u: u.is_committee_member)
)
potential_camp_officer_required = user_passes_test_improved(lambda u: u.is_potential_camp_officer)


class DataRetentionNotice(enum.Enum):
    OFFICERS = "officers"
    CAMPERS = "campers"


NamedUrl: TypeAlias = str
BreadCrumb = tuple[NamedUrl, str]


def with_breadcrumbs(breadcrumbs: list[BreadCrumb]):
    def decorator(func):
        @wraps(func)
        def wrapper(request, *args, **kwargs) -> HttpResponse:
            retval = func(request, *args, **kwargs)
            if isinstance(retval, TemplateResponse):
                retval.context_data["breadcrumbs"] = breadcrumbs
            return retval

        return wrapper

    return decorator


DATA_RETENTION_NOTICES_HTML = {
    DataRetentionNotice.OFFICERS: "cciw/officers/officer_data_retention_rules_inc.html",
    DataRetentionNotice.CAMPERS: "cciw/officers/camper_data_retention_rules_inc.html",
}

DATA_RETENTION_NOTICES_TXT = {
    DataRetentionNotice.OFFICERS: """
Share this data only with leaders or the designated CCiW officers
who assist leaders with tasks relating to officers, and no third parties.
All such people must be aware of and abide by these rules.

Keep downloaded data secure and well organised, stored only on devices that
unauthorised people do not have access to. You must be able to find and delete it later.

Delete officer addresses within 1 year of the end of the camp they
pertain to. They must be fully erased from your electronic devices and
online storage, including any copies you have made, such as attachments in
emails and backups.

""".strip(),
    DataRetentionNotice.CAMPERS: """
Share this data only with leaders and assistant leaders and no third parties.
All these people must be aware of and abide by these rules.

Keep downloaded data secure and well organised, stored only on devices that
unauthorised people do not have access to. You must be able to find and delete it later.

Delete camper information within 1 month of the end of the camp it relates to.
It must be fully erased from your electronic devices and online storage, including any
copies you have made, such as attachments in emails and backups.

""".strip(),
}

for val in DataRetentionNotice:
    assert val in DATA_RETENTION_NOTICES_HTML, f"Need to add {val} to DATA_RETENTION_NOTICES_HTML"
    assert val in DATA_RETENTION_NOTICES_TXT, f"Need to add {val} to DATA_RETENTION_NOTICES_TXT"


def show_data_retention_notice(notice_type: DataRetentionNotice, brief_title):
    """
    Decorator for downloads to show a prompt to ensure
    user knows about data retention
    """

    def decorator(func):
        @wraps(func)
        def wrapper(request, *args, **kwargs):
            htmx = "HX-Request" in request.headers
            if "data_retention_notice_seen" in request.GET:
                return func(request, *args, **kwargs)
            else:
                if htmx:
                    base_template = "cciw/officers/modal_dialog.html"
                else:
                    base_template = "cciw/officers/base.html"

                template = "cciw/officers/show_data_retention_notice.html"
                return TemplateResponse(
                    request,
                    template,
                    {
                        "base_template": base_template,
                        "include_file": DATA_RETENTION_NOTICES_HTML[notice_type],
                        "brief_title": brief_title,
                        "download_link": furl.furl(request.get_full_path()).add(
                            query_params={"data_retention_notice_seen": "1"}
                        ),
                    },
                )

        return wrapper

    return decorator


# /officers/
@staff_member_required
@never_cache
def index(request):
    """Displays a list of links/buttons for various actions."""

    # Handle redirects, since this page is LOGIN_URL
    redirect_resp = get_redirect_from_request(request)
    if redirect_resp is not None:
        return redirect_resp

    user = request.user
    context = {
        "title": "Officer home page",
    }
    context["thisyear"] = common.get_thisyear()
    context["lastyear"] = context["thisyear"] - 1
    if user.is_camp_admin or user.is_superuser:
        context["show_leader_links"] = True
        context["show_admin_link"] = True
    if user.is_cciw_secretary or user.is_superuser:
        context["show_secretary_links"] = True
        context["show_admin_link"] = True
    if user.is_dbs_officer or user.is_camp_admin or user.is_superuser:
        context["show_dbs_officer_links"] = True
    if user.is_booking_secretary or user.is_superuser:
        context["show_booking_secretary_links"] = True
    if user.is_booking_secretary or user.is_treasurer or user.is_superuser:
        context["show_booking_report_links"] = True
    if user.is_committee_member or user.is_booking_secretary or user.is_superuser:
        context["show_secretary_and_committee_links"] = True
        booking_year = most_recent_booking_year()
        if booking_year is not None:
            context["booking_stats_end_year"] = booking_year
            context["booking_stats_start_year"] = booking_year - BOOKING_STATS_PREVIOUS_YEARS

    return TemplateResponse(request, "cciw/officers/index.html", context)


officers_breadcrumbs = [("cciw-officers-index", "Officer home page")]


@staff_member_required
@camp_admin_required
@with_breadcrumbs(officers_breadcrumbs)
def leaders_index(request):
    """Displays a list of links for actions for leaders"""
    user = request.user
    thisyear = common.get_thisyear()
    show_all = "show_all" in request.GET
    camps: Iterable[Camp] = Camp.objects.all().include_other_years_info()
    if not show_all:
        camps = camps.filter(id__in=[c.id for c in user.camps_as_admin_or_leader])
    last_existing_year = Camp.objects.order_by("-year")[0].year

    return TemplateResponse(
        request,
        "cciw/officers/leaders_index.html",
        {
            "title": "Leader's tools",
            "current_camps": [c for c in camps if c.year == thisyear],
            "old_camps": [c for c in camps if c.year < thisyear],
            "statsyears": list(range(last_existing_year, last_existing_year - 3, -1)),
            "stats_end_year": last_existing_year,
            "stats_start_year": 2006,  # first year this feature existed
            "show_all": show_all,
        },
    )


leaders_breadcrumbs = officers_breadcrumbs + [("cciw-officers-leaders_index", "Leaders' tools")]


@staff_member_required
@never_cache
@with_breadcrumbs(officers_breadcrumbs)
def applications(request):
    """Displays a list of tasks related to applications."""
    user = request.user
    finished_applications = user.applications.filter(finished=True).order_by("-date_saved")
    # A NULL date_saved means they never pressed save, so there is no point
    # re-editing, so we ignore them.
    unfinished_applications = (
        user.applications.filter(finished=False).exclude(date_saved__isnull=True).order_by("-date_saved")
    )
    has_thisyears_app = thisyears_applications(user).exists()
    has_completed_app = thisyears_applications(user).filter(finished=True).exists()

    context = {
        "camps": [i.camp for i in user.invitations.filter(camp__year=common.get_thisyear())],
        "title": "Your applications",
        "finished_applications": finished_applications,
        "unfinished_applications": unfinished_applications,
        "has_thisyears_app": has_thisyears_app,
        "has_completed_app": has_completed_app,
    }

    if not has_completed_app and unfinished_applications and "edit" in request.POST:
        # Edit existing application.
        # It should now only be possible for there to be one unfinished
        # application, so we just continue with the most recent.
        return HttpResponseRedirect(reverse("admin:officers_application_change", args=(unfinished_applications[0].id,)))
    elif not has_thisyears_app and "new" in request.POST:
        # Create new application based on old one
        if finished_applications:
            new_obj = _copy_application(finished_applications[0])
        else:
            new_obj = Application.objects.create(officer=user, full_name=user.full_name)

        return HttpResponseRedirect(f"/admin/officers/application/{new_obj.id}/")

    return TemplateResponse(request, "cciw/officers/applications.html", context)


@staff_member_required
def get_application(request):
    try:
        application_id = int(request.POST["application"])
    except (KeyError, ValueError):
        raise Http404

    app = get_object_or_404(request.user.applications, id=application_id)

    format = request.POST.get("format", "")
    if format == "html":
        return HttpResponseRedirect(
            reverse("cciw-officers-view_application", kwargs=dict(application_id=application_id))
        )
    elif format == "txt":
        resp = HttpResponse(application_to_text(app), content_type="text/plain")
        resp["Content-Disposition"] = f"attachment; filename={application_txt_filename(app)}"
        return resp
    elif format == "rtf":
        resp = HttpResponse(application_to_rtf(app), content_type="text/rtf")
        resp["Content-Disposition"] = f"attachment; filename={application_rtf_filename(app)}"
        return resp
    elif format == "send":
        application_text = application_to_text(app)
        application_rtf = application_to_rtf(app)
        rtf_attachment = (application_rtf_filename(app), application_rtf, "text/rtf")

        msg = f"""Dear {request.user.first_name},

Please find attached a copy of the application you requested
 -- in plain text below and an RTF version attached.

"""
        msg = msg + application_text

        send_mail_with_attachments(
            f"[CCIW] Copy of CCiW application - {app.full_name}",
            msg,
            settings.SERVER_EMAIL,
            [formatted_email(request.user)],
            attachments=[rtf_attachment],
        )
        messages.info(request, "Email sent.")

        # Redirect back where we came from
        return HttpResponseRedirect(request.POST.get("to", "/officers/"))

    else:
        raise Http404

    return resp


@staff_member_required
def view_application_redirect(request):
    if "application_id" in request.GET:
        return HttpResponseRedirect(
            reverse("cciw-officers-view_application", kwargs=dict(application_id=request.GET["application_id"]))
        )
    raise Http404


@staff_member_required
@cache_control(max_age=3600)
@with_breadcrumbs(officers_breadcrumbs)
def view_application(request, application_id: int):
    application = get_object_or_404(Application, id=application_id)

    if application.officer_id != request.user.id and not request.user.can_manage_application_forms:
        raise PermissionDenied

    # NB, this is is called by both normal users and leaders.
    # In the latter case, request.user != app.officer

    return TemplateResponse(
        request,
        "cciw/officers/view_application.html",
        {
            "application": application,
            "officer": application.officer,
            "is_popup": True,
        },
    )


def _thisyears_camp_for_leader(user):
    leaders = list(user.people.all())
    try:
        return leaders[0].camps_as_leader.get(year=common.get_thisyear())
    except (ObjectDoesNotExist, IndexError):
        return None


@staff_member_required
@camp_admin_required
@never_cache
@with_breadcrumbs(leaders_breadcrumbs)
def manage_applications(request, camp_id: CampId):
    camp = _get_camp_or_404(camp_id)
    return TemplateResponse(
        request,
        "cciw/officers/manage_applications.html",
        {
            "title": f"Manage applications: {camp.nice_name}",
            "camp": camp,
            "finished_applications": applications_for_camp(camp).order_by("officer__first_name", "officer__last_name"),
        },
    )


def _get_camp_or_404(camp_id: CampId) -> Camp:
    try:
        return Camp.objects.by_camp_id(camp_id).get()
    except (Camp.DoesNotExist, ValueError):
        raise Http404


@staff_member_required
@camp_admin_required  # we don't care which camp they are admin for.
@never_cache
@with_breadcrumbs(leaders_breadcrumbs)
@for_htmx(use_block_from_params=True)
def manage_references(request, camp_id: CampId):
    # If referee_id is set, we just want to update part of the page.
    referee_id = request.GET.get("referee_id")
    officer = None
    officer_id = request.GET.get("officer_id")
    if officer_id is not None:
        try:
            officer = User.objects.get(id=int(officer_id))
        except (ValueError, User.DoesNotExist):
            raise Http404
    camp = _get_camp_or_404(camp_id)

    if referee_id is None:
        apps = applications_for_camp(camp, officer_ids=[officer_id] if officer is not None else None)
        app_ids = [app.id for app in apps]
        referees = Referee.objects.filter(application__in=app_ids).order_by()
    else:
        referees = Referee.objects.filter(pk=referee_id).order_by()

    referees = referees.prefetch_related(
        Prefetch("actions", queryset=ReferenceAction.objects.select_related("user"))
    ).select_related("reference", "application", "application__officer")

    all_referees = list(referees)
    if "ref_email" in request.GET:
        ref_email = request.GET["ref_email"]
        all_referees = [r for r in all_referees if r.email.lower() == ref_email.lower()]
    else:
        ref_email = None

    for referee in all_referees:
        referee.sort_key = [
            # Received come last:
            referee.reference_is_received(),
            # Not requested come first:
            referee.reference_was_requested(),
            # Then sort by:
            referee.application.officer.first_name,
            referee.application.officer.last_name,
            referee.name,
        ]
        # Note that we add this as an attribute because we also need to sort by
        # the same key client side.
        if referee.reference_is_received():
            continue  # Don't need the following
        # decorate each Reference with suggested previous References.
        add_previous_references(referee)

    all_referees.sort(key=lambda referee: referee.sort_key)

    return TemplateResponse(
        request,
        "cciw/officers/manage_references.html",
        {
            "officer": officer,
            "camp": camp,
            "title": f"Manage references: {camp.nice_name}",
            "ref_email_search": ref_email,
            "all_referees": all_referees,
        },
    )


@staff_member_required
@camp_admin_required  # we don't care which camp they are admin for.
@cache_control(max_age=3600)
def officer_history(request, officer_id: int):
    officer = get_object_or_404(User.objects.filter(id=officer_id))
    referee_pairs = [
        app.referees
        for app in (
            officer.applications.all().prefetch_related("referee_set", "referee_set__reference").order_by("-date_saved")
        )
    ]

    return TemplateResponse(
        request,
        "cciw/officers/officer_history.html",
        {
            "officer": officer,
            "referee_pairs": referee_pairs,
        },
    )


def htmx_reference_events_response(
    closeModal: bool = False,
    refreshReferee: Referee | None = None,
):
    events = {}
    if refreshReferee is not None:
        events[f"refreshReferee-{refreshReferee.id}"] = True
    if closeModal:
        events["jsCloseModal"] = closeModal

    return add_hx_trigger_header(HttpResponse(""), events)


def add_hx_trigger_header(response: HttpResponse, events: dict) -> HttpResponse:
    if events:
        response.headers["Hx-Trigger"] = json.dumps(events)
    return response


@staff_member_required
@camp_admin_required
@for_htmx(use_block_from_params=True)
def correct_referee_details(request: HttpRequest, camp_id: CampId, referee_id: int):
    referee = get_object_or_404(Referee.objects.filter(id=referee_id))
    if request.method == "POST":
        if "save" in request.POST:
            form = CorrectRefereeDetailsForm(request.POST, instance=referee)
            if form.is_valid():
                form.save()
                referee.log_details_corrected(request.user, timezone.now())
                return htmx_reference_events_response(closeModal=True, refreshReferee=referee)
        else:
            # cancel
            return htmx_reference_events_response(closeModal=True)
    else:
        form = CorrectRefereeDetailsForm(instance=referee)

    return TemplateResponse(
        request,
        "cciw/officers/correct_referee_details.html",
        {
            "form": form,
            "referee": referee,
        },
    )


def _get_previous_reference(referee: Referee, prev_ref_id: int) -> tuple[bool, Reference | None]:
    """
    Get previous reference that matches prev_ref_id, returning:
    (
       bool indicating an exact previous reference match,
       previous reference

    """
    exact_prev_reference, prev_references = get_previous_references(referee)

    if exact_prev_reference is not None:
        if exact_prev_reference.id != prev_ref_id:
            # This could happen only if the user has fiddled with URLs, or
            # there has been some update on the page.
            return (False, None)
        return (True, exact_prev_reference)
    else:
        # Get old referee data
        prev_references = [r for r in prev_references if r.id == prev_ref_id]
        if len(prev_references) != 1:
            return (False, None)
        return (False, prev_references[0])


@staff_member_required
@camp_admin_required  # we don't care which camp they are admin for.
@for_htmx(use_block_from_params=True)
def request_reference(request: HttpRequest, camp_id: CampId, referee_id: int):
    camp = _get_camp_or_404(camp_id)
    referee = get_object_or_404(Referee.objects.filter(id=referee_id))
    app = referee.application

    context = {}
    # Work out 'old_referee' or 'known_email_address', and the URL to use in the
    # message.
    try:
        prev_ref_id = int(request.GET["prev_ref_id"])
    except (KeyError, ValueError):
        prev_ref_id = None
    if prev_ref_id:
        prev_reference_is_exact, prev_reference = _get_previous_reference(referee, prev_ref_id)
        if prev_reference is None:
            return htmx_reference_events_response(closeModal=True, refreshReferee=referee)
        context["known_email_address"] = prev_reference_is_exact
        context["old_referee"] = prev_reference.referee
        url = make_ref_form_url(referee.id, prev_ref_id)
    else:
        url = make_ref_form_url(referee.id, None)
        prev_reference = None

    messageform_info = dict(
        referee=referee,
        applicant=app.officer,
        camp=camp,
        url=url,
        sender=request.user,
        update=prev_reference is not None,
    )

    if request.method == "POST":
        if "send" in request.POST:
            context["show_messageform"] = True
            form = SendReferenceRequestForm(request.POST, message_info=messageform_info)
            if form.is_valid():
                send_reference_request_email(wordwrap(form.cleaned_data["message"], 70), referee, request.user, camp)
                referee.log_request_made(request.user, timezone.now())
                return htmx_reference_events_response(closeModal=True, refreshReferee=referee)
        elif "cancel" in request.POST:
            return htmx_reference_events_response(closeModal=True)
    else:
        form = SendReferenceRequestForm(message_info=messageform_info)

    context.update(
        {
            "already_requested": referee.reference_was_requested(),
            "referee": referee,
            "app": app,
            "is_update": prev_reference is not None,
            "form": form,
        }
    )

    return TemplateResponse(request, "cciw/officers/request_reference.html", context)


@staff_member_required
@camp_admin_required
@for_htmx(use_block_from_params=True)
def fill_in_reference_manually(request: HttpRequest, camp_id: CampId, referee_id: int):
    referee = get_object_or_404(Referee.objects.filter(id=referee_id))
    reference = referee.reference if hasattr(referee, "reference") else None

    try:
        prev_ref_id = int(request.GET["prev_ref_id"])
    except (KeyError, ValueError):
        prev_ref_id = None
    if prev_ref_id:
        _, prev_reference = _get_previous_reference(referee, prev_ref_id)
    else:
        prev_reference = None

    if request.method == "POST":
        if "save" in request.POST:
            form = AdminReferenceForm(request.POST, instance=reference)
            if form.is_valid():
                form.save(referee, user=request.user)
                return htmx_reference_events_response(closeModal=True, refreshReferee=referee)
        else:
            # Cancel
            return htmx_reference_events_response(closeModal=True)
    else:
        form = get_initial_reference_form(reference, referee, prev_reference, AdminReferenceForm)

    return TemplateResponse(
        request,
        "cciw/officers/fill_in_reference_manually.html",
        {
            "referee": referee,
            "app": referee.application,
            "form": form,
            "is_update": prev_reference is not None,
        },
    )


@staff_member_required
@camp_admin_required
@for_htmx(use_block_from_params=True)
def nag_by_officer(request: HttpRequest, camp_id: CampId, referee_id: int):
    # htmx only view, runs in modal dialog
    camp = _get_camp_or_404(camp_id)
    referee = get_object_or_404(Referee.objects.filter(id=referee_id))
    app = referee.application
    officer = app.officer

    messageform_info = dict(referee=referee, officer=officer, sender=request.user, camp=camp)

    if request.method == "POST":
        if "send" in request.POST:
            form = SendNagByOfficerForm(request.POST, message_info=messageform_info)
            if form.is_valid():
                send_nag_by_officer(wordwrap(form.cleaned_data["message"], 70), officer, referee, request.user)
                referee.log_nag_made(request.user, timezone.now())
                return htmx_reference_events_response(closeModal=True, refreshReferee=referee)
        else:
            # cancel
            return htmx_reference_events_response(closeModal=True)
    else:
        form = SendNagByOfficerForm(message_info=messageform_info)

    return TemplateResponse(
        request,
        "cciw/officers/nag_by_officer.html",
        {
            "referee": referee,
            "app": app,
            "officer": officer,
            "form": form,
        },
    )


def initial_reference_form_data(referee: Referee, prev_reference: Reference | None):
    """
    Return the initial data to be used for Reference, given the current
    Referee object and the Reference object with data to be copied.
    """
    retval = {}
    if prev_reference is not None:
        # Copy data over
        for f in Reference._meta.fields:
            fname = f.attname
            if fname not in ["id", "date_created"]:
                retval[fname] = getattr(prev_reference, fname)
    retval["referee_name"] = referee.name
    return retval


def create_reference(request, referee_id: int, hash: str, prev_ref_id: int | None = None):
    """
    View for allowing referee to submit reference (create the Reference object)
    """
    context = {}
    if hash != make_ref_form_url_hash(referee_id, prev_ref_id):
        context["incorrect_url"] = True
    else:
        referee: Referee = get_object_or_404(Referee.objects.filter(id=referee_id))
        prev_reference = None
        if prev_ref_id is not None:
            prev_reference = get_object_or_404(Reference.objects.filter(id=prev_ref_id))

        if prev_reference is not None:
            context["update"] = True
            context["last_form_date"] = prev_reference.date_created if not prev_reference.inaccurate else None
            context["last_empty"] = empty_reference(prev_reference)

        reference = referee.reference if hasattr(referee, "reference") else None
        relevant_invitations = invitations_for_application(referee.application)
        role_names = sorted(list({i.role.name for i in relevant_invitations if i.role is not None}))
        context["roles"] = role_names

        if reference is not None and not empty_reference(reference):
            # It's possible that empty references have been created in the past,
            # so ensure that these don't stop people filling out form.
            context["already_submitted"] = True
        else:
            if request.method == "POST":
                form = ReferenceForm(request.POST, instance=reference)
                if form.is_valid():
                    form.save(referee)
                    return HttpResponseRedirect(reverse("cciw-officers-create_reference_thanks"))
            else:
                form = get_initial_reference_form(reference, referee, prev_reference, ReferenceForm)
            context["form"] = form
        context["officer"] = referee.application.officer
    return TemplateResponse(request, "cciw/officers/create_reference.html", context)


def get_initial_reference_form(reference: Reference, referee: Referee, prev_reference: Reference | None, form_class):
    initial_data = initial_reference_form_data(referee, prev_reference)
    if reference is not None:
        # For the case where a Reference has been created (accidentally)
        # by an admin, we need to re-use it, rather than create another.
        if empty_reference(reference):
            # Need to fill data
            for k, v in initial_data.items():
                setattr(reference, k, v)
        form = form_class(instance=reference)
    else:
        form = form_class(initial=initial_data)
    return form


def create_reference_thanks(request):
    return TemplateResponse(request, "cciw/officers/create_reference_thanks.html", {})


@staff_member_required
@camp_admin_required
@cache_control(max_age=3600)
def view_reference(request, reference_id: int):
    reference = get_object_or_404(Reference.objects.filter(id=reference_id))
    return TemplateResponse(
        request,
        "cciw/officers/view_reference_form.html",
        {
            "reference": reference,
            "officer": reference.referee.application.officer,
            "referee": reference.referee,
            "is_popup": True,
        },
    )


@staff_member_required
@camp_admin_required
@for_htmx(use_block_from_params=True)
@with_breadcrumbs(leaders_breadcrumbs)
def officer_list(
    request: HttpRequest,
    camp_id: CampId,
) -> HttpResponse:
    return _officer_list(request, camp_id)


# undecorated for internal redirect use
def _officer_list(
    request: HttpRequest,
    camp_id: CampId,
    *,
    selected_officers: set[User] | None = None,
    open_chooseofficers: bool | None = None,
    search_query: str = "",
    selected_role: int | None = None,
    add_officer_message: str = "",
) -> TemplateResponse:
    camp = _get_camp_or_404(camp_id)
    officer_list = OfficerList(camp)
    camp_roles = CampRole.objects.all()

    try:
        # From create_officer view
        created_officer = User.objects.get(id=int(request.GET.get("created_officer_id", "")))
    except (ValueError, User.DoesNotExist):
        created_officer = None

    selected_officers = selected_officers or set()

    if request.method == "POST":
        # "Add officer" functionality
        chosen_officers = [
            officer for officer in officer_list.addable_officers if f"chooseofficer_{officer.id}" in request.POST
        ]
        add_previous_role = "add_previous_role" in request.POST
        add_new_role = "add_new_role" in request.POST
        if add_previous_role or add_new_role:
            if add_new_role:
                new_role = CampRole.objects.get(id=int(request.POST["new_role"]))
                added_officers = [add_officer_to_camp(camp, o, new_role) for o in chosen_officers]
            elif add_previous_role:
                added_officers = [
                    add_officer_to_camp(camp, o, role)
                    for o in chosen_officers
                    if (role := officer_list.get_previous_role(o)) is not None
                ]
            else:
                added_officers = []
            # if we successfully process, we remove from chosen_officer_ids,
            # but preserve the state of checkboxes we weren't able to handle.
            selected_officers = set(chosen_officers) - set(added_officers)
            if selected_officers:
                add_officer_message = "Some officers could not be added because their previous role is not known"

        # "Remove officer" functionality
        if "remove" in request.POST:
            remove_officer_from_camp(camp, User.objects.get(id=int(request.POST["officer_id"])))

        # Internal redirect, to refresh data from DB
        return _officer_list(
            make_get_request(request),
            camp_id=camp.url_id,
            # Propagate some state from POST:
            selected_officers=selected_officers,
            open_chooseofficers=bool(chosen_officers),  # if we just added some, probably want to add more
            search_query=request.POST.get("search", ""),
            selected_role=int(request.POST["new_role"]) if ("new_role" in request.POST) else None,
            add_officer_message=add_officer_message,
        )

    # Should the 'choose officers' component default to open?
    if open_chooseofficers is None:
        open_chooseofficers = (
            len(officer_list.invitations) == 0  # no officers added yet
            or created_officer is not None  # just created one in system, will want to add them
        )
    context = {
        "camp": camp,
        "title": f"Officer list: {camp.nice_name}, {camp.leaders_formatted}",
        "invitations": officer_list.invitations,
        "candidate_officers": officer_list.candidate_officers,
        "open_chooseofficers": open_chooseofficers,
        "selected_officers": selected_officers,
        "add_officer_message": add_officer_message,
        "camp_roles": camp_roles,
        "selected_role": selected_role,
        "address_all": address_for_camp_officers(camp),
        "created_officer": created_officer,
        "search_query": search_query,
    }

    return TemplateResponse(request, "cciw/officers/officer_list.html", context)


@staff_member_required
@camp_admin_required
def update_officer(request):
    # Partial page, via htmx
    invitation = Invitation.objects.select_related("role", "officer").get(id=int(request.GET["invitation_id"]))
    officer = invitation.officer
    mode = "edit"
    if request.method == "POST":
        if "save" in request.POST:
            form = UpdateOfficerForm(data=request.POST, instance=officer, invitation=invitation)
            if form.is_valid():
                form.save()
                mode = "display"
        else:
            # Cancel
            mode = "display"
            form = None
    else:
        form = UpdateOfficerForm(instance=officer, invitation=invitation)

    return TemplateResponse(
        request,
        "cciw/officers/officer_list_officer_row_inc.html",
        {
            "mode": mode,
            "form": form,
            "invitation": invitation,
        },
    )


@staff_member_required
@camp_admin_required
@with_breadcrumbs(leaders_breadcrumbs)
def officer_application_status(request, camp_id: CampId):
    camp = _get_camp_or_404(camp_id)
    return TemplateResponse(
        request,
        "cciw/officers/officer_application_status.html",
        {
            "camp": camp,
            "title": f"Application form status: {camp.nice_name}",
            "officers_noapplicationform": camp_slacker_list(camp),
            "address_noapplicationform": address_for_camp_slackers(camp),
            "officers_serious_slackers": camp_serious_slacker_list(camp),
        },
    )


def correct_email(request):
    context = {}
    try:
        username, new_email = signing.loads(
            request.GET.get("t", ""), salt="cciw-officers-correct_email", max_age=60 * 60 * 24 * 10
        )  # 10 days
    except signing.BadSignature:
        context["message"] = (
            "The URL was invalid. Please ensure you copied the URL from the email correctly, "
            "or contact the webmaster if you are having difficulties"
        )
    else:
        u = get_object_or_404(User.objects.filter(username=username))
        u.email = new_email
        u.save()
        context["message"] = "Your email address has been updated, thanks."
        context["success"] = True

    return TemplateResponse(request, "cciw/officers/email_update.html", context)


def correct_application(request):
    context = {}
    try:
        application_id, email = signing.loads(
            request.GET.get("t", ""), salt="cciw-officers-correct_application", max_age=60 * 60 * 24 * 10
        )  # 10 days
    except signing.BadSignature:
        context["message"] = (
            "The URL was invalid. Please ensure you copied the URL from the email correctly, "
            "or contact the webmaster if you are having difficulties."
        )
    else:
        application = get_object_or_404(Application.objects.filter(id=application_id))
        application.address_email = email
        application.save()
        context["message"] = "Your application form email address has been updated, thanks."
        context["success"] = True

    return TemplateResponse(request, "cciw/officers/email_update.html", context)


@staff_member_required
@camp_admin_required
@with_breadcrumbs(leaders_breadcrumbs)
def create_officer(request):
    duplicate_message, allow_confirm, existing_users = "", True, []
    message = ""
    if request.method == "POST":
        form = CreateOfficerForm(request.POST)
        process_form = False
        if form.is_valid():
            duplicate_message, allow_confirm, existing_users = form.check_duplicates()
            if "add" in request.POST:
                # If no duplicates, we can process without confirmation
                if not duplicate_message:
                    process_form = True

            elif allow_confirm and "confirm" in request.POST:
                process_form = True

            if process_form:
                u = form.save()
                redirect_resp = get_redirect_from_request(request)
                if redirect_resp:
                    redirect_url = redirect_resp["Location"]
                    message = f"Officer {u.full_name} has been added to the system and emailed."
                    match: ResolverMatch = get_resolver().resolve(redirect_url)
                    if match is not None and match.func == officer_list:
                        message += " Don't forget to choose a role and add them to your officer list!"

                    redirect_resp["Location"] = furl.furl(redirect_url).add({"created_officer_id": u.id}).url
                    messages.info(request, message)
                    return redirect_resp
                else:
                    messages.info(
                        request,
                        f"Officer {u.full_name} has been added and emailed.  You can add another if required.",
                    )
                return HttpResponseRedirect(".")

    else:
        form = CreateOfficerForm()

    return TemplateResponse(
        request,
        "cciw/officers/create_officer.html",
        {
            "title": "Add officer to system",
            "form": form,
            "duplicate_message": duplicate_message,
            "existing_users": existing_users,
            "allow_confirm": allow_confirm,
            "message": message,
        },
    )


@staff_member_required
@camp_admin_required
@require_POST
def resend_email(request):
    officer_id = int(request.POST["officer_id"])
    user = User.objects.get(pk=officer_id)
    create.email_officer(user, update=True)
    return TemplateResponse(
        request,
        "cciw/officers/resend_email_form_inc.html",
        {
            "officer_id": officer_id,
            "caption": "Sent!",
        },
    )


@staff_member_required
@camp_admin_required
@show_data_retention_notice(DataRetentionNotice.OFFICERS, "Officer data")
def export_officer_data(request, camp_id: CampId):
    camp = _get_camp_or_404(camp_id)
    return spreadsheet_response(
        officer_data_to_spreadsheet(camp),
        f"CCIW-camp-{camp.url_id}-officers",
        notice=DataRetentionNotice.OFFICERS,
    )


@staff_member_required
@camp_admin_required
@show_data_retention_notice(DataRetentionNotice.CAMPERS, "Camper data")
def export_camper_data(request, camp_id: CampId):
    camp = _get_camp_or_404(camp_id)
    return spreadsheet_response(
        camp_bookings_to_spreadsheet(camp),
        f"CCIW-camp-{camp.url_id}-campers",
        notice=DataRetentionNotice.CAMPERS,
    )


@staff_member_required
@booking_secretary_required
@show_data_retention_notice(DataRetentionNotice.CAMPERS, "Camper data")
def export_camper_data_for_year(request, year: int):
    return spreadsheet_response(
        year_bookings_to_spreadsheet(year),
        f"CCIW-bookings-{year}",
        notice=DataRetentionNotice.CAMPERS,
    )


@staff_member_required
@camp_admin_required
@show_data_retention_notice(DataRetentionNotice.CAMPERS, "Camper sharable transport details")
def export_sharable_transport_details(request, camp_id: CampId):
    camp = _get_camp_or_404(camp_id)
    return spreadsheet_response(
        camp_sharable_transport_details_to_spreadsheet(camp),
        f"CCIW-camp-{camp.url_id}-transport-details",
        notice=DataRetentionNotice.CAMPERS,
    )


@staff_member_required
@potential_camp_officer_required
def officer_files(request, path: str):
    return get_protected_download("officers", path)


@staff_member_required
@camp_admin_required
@with_breadcrumbs(leaders_breadcrumbs)
def officer_stats(request, year: int):
    camps = list(Camp.objects.filter(year=year).order_by("camp_name__slug"))
    if len(camps) == 0:
        raise Http404

    charts = []
    for camp in camps:
        df = get_camp_officer_stats(camp)
        df["References ÷ 2"] = df["References"] / 2  # Make it match the height of others
        df.pop("References")
        charts.append(
            (
                camp,
                pandas_highcharts.core.serialize(
                    df, title=f"{camp.name} - {camp.leaders_formatted}", output_type="json"
                ),
            )
        )
    return TemplateResponse(
        request,
        "cciw/officers/stats.html",
        {
            "camps": camps,
            "title": f"Officer stats {year}",
            "year": year,
            "charts": charts,
        },
    )


@staff_member_required
@camp_admin_required
@with_breadcrumbs(leaders_breadcrumbs)
def officer_stats_trend(request, start_year: int, end_year: int):
    start_year = int(start_year)
    end_year = int(end_year)
    data = get_camp_officer_stats_trend(start_year, end_year)
    for c in data.columns:
        if "fraction" not in c:
            data.pop(c)
    fraction_to_percent(data)
    return TemplateResponse(
        request,
        "cciw/officers/stats_trend.html",
        {
            "title": f"Officer stats {start_year}-{end_year}",
            "start_year": start_year,
            "end_year": end_year,
            "chart_data": pandas_highcharts.core.serialize(
                data, title=f"Officer stats {start_year} - {end_year}", output_type="json"
            ),
        },
    )


def fraction_to_percent(data):
    for col_name in list(data.columns):
        parts = col_name.split(" ")
        new_name = " ".join("%" if p.lower() == "fraction" else p for p in parts)
        if new_name != col_name:
            data[new_name] = data[col_name] * 100
            data.pop(col_name)


@staff_member_required
@camp_admin_required
def officer_stats_download(request, year: int) -> HttpResponse:
    camps = list(Camp.objects.filter(year=year).order_by("camp_name__slug"))
    builder = ExcelFromDataFrameBuilder()
    for camp in camps:
        builder.add_sheet_from_dataframe(str(camp.url_id), get_camp_officer_stats(camp))
    return spreadsheet_response(
        builder,
        f"CCIW-officer-stats-{year}",
        notice=None,
    )


@staff_member_required
@camp_admin_required
def officer_stats_trend_download(request, start_year: int, end_year: int) -> HttpResponse:
    builder = ExcelFromDataFrameBuilder()
    builder.add_sheet_from_dataframe("Officer stats trend", get_camp_officer_stats_trend(start_year, end_year))
    return spreadsheet_response(builder, f"CCIW-officer-stats-trend-{start_year}-{end_year}", notice=None)


@staff_member_required
@dbs_officer_or_camp_admin_required
@ensure_csrf_cookie
@with_breadcrumbs(officers_breadcrumbs)
@for_htmx(use_block_from_params=True)
def manage_dbss(request, year: int) -> HttpResponse:
    # We need a lot of information. Try to get it in a few up-front queries
    camps = list(Camp.objects.filter(year=year).order_by("camp_name__slug"))
    if len(camps) == 0:
        raise Http404

    # Selected camps:
    # We need to support URLs that indicate which camp to select, so we
    # can permalink nicely.
    if "camp" in request.GET:
        selected_camp_slugs = set(request.GET.getlist("camp"))
        selected_camps = {c for c in camps if c.slug_name in selected_camp_slugs}
    else:
        if "Hx-Request" in request.headers and request.GET.get("use_block", "") == "content":
            # They deselected the last checkbox, we should show them nothing for UI consistency.
            # In other cases (e.g. use_block = table-body, officer_id=...), we should include all camps.
            selected_camps = set()
        else:
            # Assume all, because having none is never useful
            selected_camps = set(camps)

    try:
        officer_id = int(request.GET["officer_id"])
    except (KeyError, ValueError):
        officer_id = None

    officers_and_dbs_info = get_officers_with_dbs_info_for_camps(camps, selected_camps, officer_id=officer_id)

    return TemplateResponse(
        request,
        "cciw/officers/manage_dbss.html",
        {
            "title": f"Manage DBSs {year}",
            "officers_and_dbs_info": officers_and_dbs_info,
            "camps": camps,
            "selected_camps": selected_camps,
            "year": year,
            "CheckType": DBSCheck.CheckType,
            "external_dbs_officer": settings.EXTERNAL_DBS_OFFICER,
        },
    )


def htmx_dbs_events_response(
    closeModal: bool = False,
    refreshOfficer: User | None = None,
) -> HttpResponse:
    events = {}
    if refreshOfficer is not None:
        events[f"refreshOfficer-{refreshOfficer.id}"] = True
    if closeModal:
        events["jsCloseModal"] = closeModal

    return add_hx_trigger_header(HttpResponse(""), events)


@staff_member_required
@dbs_officer_required
def mark_dbs_sent(request):
    officer_id = int(request.POST["officer_id"])
    officer = User.objects.get(id=officer_id)
    if "mark_sent" in request.POST:
        request.user.dbsactions_performed.create(officer=officer, action_type=DBSActionLogType.FORM_SENT)
    elif "undo_last_mark_sent" in request.POST:
        request.user.dbsactions_performed.remove_last(officer=officer, action_type=DBSActionLogType.FORM_SENT)

    return htmx_dbs_events_response(refreshOfficer=officer)


def modal_dialog_message_form(
    request,
    context,
    *,
    template_name: str,
    messageform_info: dict,
    send_email: Callable[[str], None],
    messageform_class=type,
    success_response: HttpResponse,
    cancel_response: HttpResponse,
):
    if request.method == "POST":
        if "send" in request.POST:
            messageform = messageform_class(request.POST, message_info=messageform_info)
            if messageform.is_valid():
                send_email(wordwrap(messageform.cleaned_data["message"], 70))
                return success_response
        else:
            return cancel_response
    else:
        messageform = messageform_class(message_info=messageform_info)

    context["messageform"] = messageform
    return TemplateResponse(request, template_name, context)


@staff_member_required
@dbs_officer_required
@for_htmx(use_block_from_params=True)
def dbs_consent_alert_leaders(request, application_id: int):
    app = get_object_or_404(Application.objects.filter(id=application_id))
    officer = app.officer
    camps = camps_for_application(app)
    context = {"officer": officer}
    messageform_info = {
        "application": app,
        "officer": officer,
        "camps": camps,
        "domain": common.get_current_domain(),
        "sender": request.user,
    }

    def send_email(message):
        send_dbs_consent_alert_leaders_email(message, officer, camps)
        request.user.dbsactions_performed.create(officer=officer, action_type=DBSActionLogType.LEADER_ALERT_SENT)

    return modal_dialog_message_form(
        request,
        context,
        template_name="cciw/officers/dbs_consent_alert_leaders.html",
        messageform_info=messageform_info,
        messageform_class=DbsConsentProblemForm,
        send_email=send_email,
        cancel_response=htmx_dbs_events_response(closeModal=True),
        success_response=htmx_dbs_events_response(closeModal=True, refreshOfficer=officer),
    )


@staff_member_required
@dbs_officer_required
def request_dbs_form_action(request, application_id: int):
    app = get_object_or_404(Application.objects.filter(id=application_id))
    external_dbs_officer = settings.EXTERNAL_DBS_OFFICER
    officer = app.officer
    context = {
        "officer": officer,
        "external_dbs_officer": external_dbs_officer,
    }
    messageform_info = {
        "external_dbs_officer": external_dbs_officer,
        "application": app,
        "officer": officer,
        "sender": request.user,
    }

    def send_email(message):
        send_request_for_dbs_form_email(message, officer, request.user)
        request.user.dbsactions_performed.create(
            officer=officer, action_type=DBSActionLogType.REQUEST_FOR_DBS_FORM_SENT
        )

    return modal_dialog_message_form(
        request,
        context,
        template_name="cciw/officers/request_dbs_form_action.html",
        messageform_info=messageform_info,
        messageform_class=RequestDbsFormForm,
        send_email=send_email,
        cancel_response=htmx_dbs_events_response(closeModal=True),
        success_response=htmx_dbs_events_response(closeModal=True, refreshOfficer=officer),
    )


@staff_member_required
@dbs_officer_required
@for_htmx(use_block_from_params=True)
def dbs_checked_online(request: HttpRequest, officer_id: int):
    officer = User.objects.get(id=officer_id)
    dbs_number = request.GET.get("dbs_number", "")
    old_dbs_check = officer.dbs_checks.filter(dbs_number=dbs_number).order_by("-completed").first()
    form_initial = {
        "dbs_number": dbs_number,
        "registered_with_dbs_update": True,
        "completed": date.today(),
        "check_type": DBSCheck.CheckType.ONLINE,
    }
    if old_dbs_check:
        form_initial.update(
            {
                "requested_by": old_dbs_check.requested_by,
                "other_organisation": old_dbs_check.other_organisation,
            }
        )
    return _dbscheck_create_form(request, officer, form_initial)


@staff_member_required
@dbs_officer_required
@for_htmx(use_block_from_params=True)
def dbs_register_received(request: HttpRequest, officer_id: int):
    officer = User.objects.get(id=officer_id)
    form_initial = {
        "check_type": DBSCheck.CheckType.FORM,
    }
    return _dbscheck_create_form(request, officer, form_initial)


def _dbscheck_create_form(request: HttpRequest, officer: User, form_initial: dict):
    dbscheck_admin_instance = admin_site._registry[DBSCheck]
    if request.method == "POST":
        if "save" in request.POST:
            form = DBSCheckForm(request.POST)
            if form.is_valid():
                dbs_check = form.save(officer=officer)
                # Copy what admin does for addition action
                dbscheck_admin_instance.log_addition(request, dbs_check, [{"added": {}}])
                # TODO add admin.LogEntry
                return htmx_dbs_events_response(refreshOfficer=officer, closeModal=True)
        else:
            return htmx_dbs_events_response(closeModal=True)
    else:
        if "completed" in form_initial and not isinstance(form_initial["completed"], str):
            form_initial["completed"] = form_initial["completed"].strftime("%Y-%m-%d")
        form = DBSCheckForm(initial=form_initial)

    return TemplateResponse(
        request,
        "cciw/officers/add_dbs_check.html",
        {
            "officer": officer,
            "form": form,
        },
    )


@staff_member_required
@with_breadcrumbs(officers_breadcrumbs)
def officer_info(request):
    return TemplateResponse(
        request,
        "cciw/officers/info.html",
        {
            "title": "Information for officers",
            "show_wiki_link": request.user.is_wiki_user,
        },
    )


# treasurer gets to see these to know how much money
# to transfer to camp leaders.
@booking_secretary_or_treasurer_required
@with_breadcrumbs(officers_breadcrumbs)
def booking_secretary_reports(request, year: int):
    from cciw.bookings.models import Booking, booking_report_by_camp, outstanding_bookings_with_fees

    # 1. Camps and their booking levels.
    camps = booking_report_by_camp(year)

    # 2. Online bookings needing attention
    to_approve = Booking.objects.need_approving().for_year(year)

    # 3. Fees
    outstanding = outstanding_bookings_with_fees(year)

    export_start = datetime(year - 1, 11, 1)  # November previous year
    export_end = datetime(year, 10, 31)  # November this year
    export_data_link = reverse("cciw-officers-export_payment_data") + "?start={start}&end={end}".format(
        start=export_start.strftime(EXPORT_PAYMENT_DATE_FORMAT), end=export_end.strftime(EXPORT_PAYMENT_DATE_FORMAT)
    )

    return TemplateResponse(
        request,
        "cciw/officers/booking_secretary_reports.html",
        {
            "title": f"Bookings {year}",
            "year": year,
            "stats_start_year": year - BOOKING_STATS_PREVIOUS_YEARS,
            "camps": camps,
            "bookings": outstanding,
            "to_approve": to_approve,
            "export_start": export_start,
            "export_end": export_end,
            "export_data_link": export_data_link,
        },
    )


@booking_secretary_required
def export_payment_data(request):
    date_start = request.GET["start"]
    date_end = request.GET["end"]
    date_start = datetime.strptime(date_start, EXPORT_PAYMENT_DATE_FORMAT).replace(
        tzinfo=timezone.get_default_timezone()
    )
    date_end = datetime.strptime(date_end, EXPORT_PAYMENT_DATE_FORMAT).replace(tzinfo=timezone.get_default_timezone())
    return spreadsheet_response(
        payments_to_spreadsheet(date_start, date_end),
        f"CCIW-payments-{date_start:%Y-%m-%d}-to-{date_end:%Y-%m-%d}",
        notice=DataRetentionNotice.CAMPERS,
    )


def _parse_year_or_camp_ids(start_year, end_year, camp_ids):
    if camp_ids is not None:
        return None, None, [_get_camp_or_404(camp_id) for camp_id in camp_ids]
    else:
        return int(start_year), int(end_year), None


def _get_booking_progress_stats_from_params(start_year, end_year, camp_ids, **kwargs):
    start_year, end_year, camps = _parse_year_or_camp_ids(start_year, end_year, camp_ids)
    if camps is not None:
        data_dates, data_rel_days = get_booking_progress_stats(camps=camps, **kwargs)
    else:
        data_dates, data_rel_days = get_booking_progress_stats(start_year=start_year, end_year=end_year, **kwargs)

    return start_year, end_year, camps, data_dates, data_rel_days


@staff_member_required
@camp_admin_required
@with_breadcrumbs(officers_breadcrumbs)
def booking_progress_stats(request, start_year: int = None, end_year: int = None, camp_ids: list[CampId] = None):
    start_year, end_year, camp_objs, data_dates, data_rel_days = _get_booking_progress_stats_from_params(
        start_year, end_year, camp_ids, overlay_years=True
    )
    return TemplateResponse(
        request,
        "cciw/officers/booking_progress_stats.html",
        {
            "title": "Booking progress" + (f" {start_year}-{end_year}" if start_year else ""),
            "start_year": start_year,
            "end_year": end_year,
            "camps": camp_objs,
            "camp_ids": camp_ids,
            "dates_chart_data": pandas_highcharts.core.serialize(
                data_dates, title="Bookings by date", output_type="json"
            ),
            "rel_days_chart_data": pandas_highcharts.core.serialize(
                data_rel_days,
                title="Bookings by days relative to start of camp",
                output_type="json",
            ),
        },
    )


@staff_member_required
@camp_admin_required
def booking_progress_stats_download(
    request, start_year: int = None, end_year: int = None, camp_ids: list[CampId] = None
):
    start_year, end_year, camp_objs, data_dates, data_rel_days = _get_booking_progress_stats_from_params(
        start_year, end_year, camp_ids, overlay_years=False
    )
    builder = ExcelFromDataFrameBuilder()
    builder.add_sheet_from_dataframe("Bookings against date", data_dates)
    builder.add_sheet_from_dataframe("Days relative to start of camp", data_rel_days)
    if camp_ids is not None:
        filename = f"CCIW-booking-progress-stats-{'_'.join(str(camp_id) for camp_id in camp_ids)}"
    else:
        filename = f"CCIW-booking-progress-stats-{start_year}-{end_year}"
    return spreadsheet_response(
        builder,
        filename,
        notice=None,
    )


@staff_member_required
@secretary_or_committee_required
@with_breadcrumbs(officers_breadcrumbs)
def booking_summary_stats(request, start_year: int, end_year: int):
    chart_data = get_booking_summary_stats(start_year, end_year)
    chart_data.pop("Total")
    return TemplateResponse(
        request,
        "cciw/officers/booking_summary_stats.html",
        {
            "title": f"Booking summary {start_year}-{end_year}",
            "start_year": start_year,
            "end_year": end_year,
            "chart_data": pandas_highcharts.core.serialize(chart_data, output_type="json"),
        },
    )


@staff_member_required
@secretary_or_committee_required
def booking_summary_stats_download(request, start_year: int, end_year: int):
    data = get_booking_summary_stats(start_year, end_year)
    builder = ExcelFromDataFrameBuilder()
    builder.add_sheet_from_dataframe("Bookings", data)
    return spreadsheet_response(builder, f"CCIW-booking-summary-stats-{start_year}-{end_year}", notice=None)


def _get_booking_ages_stats_from_params(start_year, end_year, camp_ids) -> tuple[int, int, list[Camp], pd.DataFrame]:
    start_year, end_year, camps = _parse_year_or_camp_ids(start_year, end_year, camp_ids)
    if camps is not None:
        data = get_booking_ages_stats(camps=camps, include_total=True)
    else:
        data = get_booking_ages_stats(start_year=start_year, end_year=end_year, include_total=False)
    return start_year, end_year, camps, data


@staff_member_required
@camp_admin_required
@with_breadcrumbs(officers_breadcrumbs)
def booking_ages_stats(
    request, start_year: int = None, end_year: int = None, camp_ids: list[CampId] = None, single_year: int = None
):
    if single_year is not None:
        camps = Camp.objects.filter(year=int(single_year))
        return HttpResponseRedirect(
            reverse("cciw-officers-booking_ages_stats_custom", kwargs={"camp_ids": [c.url_id for c in camps]})
        )
    start_year, end_year, camps, data = _get_booking_ages_stats_from_params(start_year, end_year, camp_ids)
    if "Total" in data:
        data.pop("Total")

    if camps:
        if all(c.year == camps[0].year for c in camps):
            stack_columns = True
        else:
            stack_columns = False
    else:
        stack_columns = False

    # Use colors defined for camps if possible. To get them to line up with data
    # series, we have to sort in the same way the pandas_highcharts does i.e. by
    # series name
    colors = []
    if camps:
        colors = [color for (title, color) in sorted((str(c.url_id), c.camp_name.color) for c in camps)]
        if len(set(colors)) != len(colors):
            # Not enough - fall back to auto
            colors = []

    return TemplateResponse(
        request,
        "cciw/officers/booking_ages_stats.html",
        {
            "title": "Camper ages stats" + (f" {start_year}-{end_year}" if start_year else ""),
            "start_year": start_year,
            "end_year": end_year,
            "camps": camps,
            "camp_ids": camp_ids,
            "chart_data": pandas_highcharts.core.serialize(data, title="Age of campers", output_type="json"),
            "colors_data": colors,
            "stack_columns": stack_columns,
        },
    )


@staff_member_required
@camp_admin_required
def booking_ages_stats_download(request, start_year: int = None, end_year: int = None, camp_ids: list[CampId] = None):
    start_year, end_year, camps, data = _get_booking_ages_stats_from_params(start_year, end_year, camp_ids)
    builder = ExcelFromDataFrameBuilder()
    builder.add_sheet_from_dataframe("Age of campers", data)
    if camp_ids is not None:
        filename = f"CCIW-booking-ages-stats-{'_'.join(str(camp_id) for camp_id in camp_ids)}"
    else:
        filename = f"CCIW-booking-ages-stats-{start_year}-{end_year}"
    return spreadsheet_response(builder, filename, notice=None)


@cciw_secretary_or_booking_secretary_required
def brochure_mailing_list(request, year: int):
    return spreadsheet_response(
        addresses_for_mailing_list(year), f"CCIW-mailing-list-{year}", notice=DataRetentionNotice.CAMPERS
    )


def spreadsheet_response(
    builder: ExcelBuilder,
    filename: str,
    *,
    notice: DataRetentionNotice | None,
) -> HttpResponse:
    output = builder.to_bytes()

    if notice is not None:
        workbook: openpyxl.Workbook = xl.workbook_from_bytes(builder.to_bytes())
        sheet = workbook.create_sheet("Notice", 0)
        c_header = sheet.cell(1, 1)
        c_header.value = "Data retention notice:"
        c_header.font = xl.header_font

        for row_idx, line in enumerate(notice_to_lines(notice), start=3):
            c = sheet.cell(row_idx, 1)
            c.value = line
            c.font = xl.default_font
        sheet.column_dimensions["A"].width = 100

        output = xl.workbook_to_bytes(workbook)
    response = HttpResponse(output, content_type=builder.mimetype)
    response["Content-Disposition"] = f"attachment; filename={filename}.{builder.file_ext}"
    return response


def notice_to_lines(notice: DataRetentionNotice) -> list[str]:
    txt = DATA_RETENTION_NOTICES_TXT[notice]
    return list(txt.split("\n"))


@booking_secretary_required
@json_response
def place_availability_json(request):
    retval = {"status": "success"}
    camp_id = int(request.GET["camp_id"])
    camp: Camp = Camp.objects.get(id=camp_id)
    places = camp.get_places_left()
    retval["result"] = dict(total=places.total, male=places.male, female=places.female)
    return retval


@booking_secretary_required
@json_response
def booking_problems_json(request):
    """
    Get the booking problems associated with the data POSTed.
    """
    # This is used by the admin.
    # We have to create a Booking object, but not save it.
    from cciw.bookings.admin import BookingAdminForm

    # Make it easy on front end:
    data = request.POST.copy()
    with contextlib.suppress(KeyError):
        data["created_at"] = data["created_at_0"] + " " + data["created_at_1"]

    if "booking_id" in data:
        booking_obj = Booking.objects.get(id=int(data["booking_id"]))
        if "created_online" not in data:
            # readonly field, data not included in form
            data["created_online"] = booking_obj.created_online
        form = BookingAdminForm(data, instance=booking_obj)
    else:
        form = BookingAdminForm(data)

    retval = {"status": "success"}
    if form.is_valid():
        retval["valid"] = True
        instance: Booking = form.save(commit=False)
        # We will get errors later on if prices don't exist for the year chosen, so
        # we check that first.
        if not is_booking_open(instance.camp.year):
            retval["problems"] = [f"Prices have not been set for the year {instance.camp.year}"]
        else:
            problems, warnings = instance.get_booking_problems(booking_sec=True)
            retval["problems"] = problems + warnings
    else:
        retval["valid"] = False
        retval["errors"] = form.errors
    return retval


@json_response
@staff_member_required
@booking_secretary_required
def get_booking_expected_amount_due(request):
    fail = {"status": "success", "amount": None}
    try:
        # If we use a form to construct an object, we won't get pass
        # validation. So we construct a partial object, doing manual parsing of
        # posted vars.

        if "id" in request.POST:
            # Start with saved data if it is available
            b = Booking.objects.get(id=int(request.POST["id"]))
        else:
            b = Booking()
        b.price_type = int(request.POST["price_type"])
        b.camp_id = int(request.POST["camp"])
        b.early_bird_discount = "early_bird_discount" in request.POST
        b.state = int(request.POST["state"])
    except (ValueError, KeyError):  # not a valid price_type/camp, data missing
        return fail
    try:
        amount = b.expected_amount_due()
    except Price.DoesNotExist:
        return fail

    if amount is not None:
        amount = str(amount)  # convert decimal
    return {"status": "success", "amount": amount}


cciw_password_reset = PasswordResetView.as_view(form_class=CciwPasswordResetForm)
